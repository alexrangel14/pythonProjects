# openpyxl gives us methods to manipulate excel files
import openpyxl as xl
# Functionality to add a bar chart
from openpyxl.chart import BarChart, Reference

# Function that takes in a excel file and "corrects the price" by multiplying by .9 and
# adding this correction in a new column

def process_workbook(filename):
    # load_workbook allows us to load an existing excel file into program
    # that we can then manipulate
    wb = xl.load_workbook(filename)
    # Select the sheet in excel you want to manipulate
    sheet = wb['Sheet1']
    # Go through every row in the sheet
    for row in range(2,sheet.max_row + 1):
        # Get the specific cells info
        cell = sheet.cell(row,3)
        # Get new modified value from the cell
        correctedPrice = cell.value * .9
        # Get the cell at a new column
        corrected_price_cell = sheet.cell(row,4)
        # Set the value for that new cell column location
        corrected_price_cell.value = correctedPrice
    # Reference get a cell range from the sheet
    values = Reference(sheet, min_row=2, max_row=sheet.max_row, min_col=4, max_col=4)
    # Create an empty BarChart() object
    chart = BarChart()
    # Add the data into the bar chart
    chart.add_data(values)
    # Add the bar char at location to the sheet
    sheet.add_chart(chart,'e2')
    # Save changes made to the workbook back to excel file
    wb.save(filename)